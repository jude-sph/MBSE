import json
from pathlib import Path
from src.exporter import export_json, export_xlsx, export_text
from src.models import MBSEModel, Meta, Requirement, Link


def make_test_model():
    return MBSEModel(
        meta=Meta(source_file="test.xlsx", mode="capella", selected_layers=["operational_analysis"],
                  llm_provider="openrouter", llm_model="test"),
        requirements=[Requirement(id="REQ-001", text="Test requirement", source_dig="DIG-001")],
        layers={"operational_analysis": {
            "entities": [{"id": "OE-001", "name": "PIB Icebreaker", "type": "OperationalEntity", "actors": ["CO"]}],
            "capabilities": [{"id": "OC-001", "name": "Conduct SAR", "involved_entities": ["OE-001"]}],
            "scenarios": [],
            "activities": [],
        }},
        links=[Link(id="LNK-001", source="OE-001", target="REQ-001", type="satisfies", description="test link")],
        instructions={"tool": "Capella 7.0", "steps": [
            {"step": 1, "action": "Create project", "detail": "File > New > Capella Project", "layer": "general"}
        ]},
    )


def test_export_json(tmp_path):
    model = make_test_model()
    path = export_json(model, tmp_path / "output.json")
    assert path.exists()
    data = json.loads(path.read_text())
    assert data["meta"]["mode"] == "capella"
    assert len(data["requirements"]) == 1
    assert len(data["links"]) == 1


def test_export_xlsx(tmp_path):
    model = make_test_model()
    path = export_xlsx(model, tmp_path / "output.xlsx")
    assert path.exists()
    import openpyxl
    wb = openpyxl.load_workbook(path)
    assert "Links" in wb.sheetnames
    assert "Instructions" in wb.sheetnames


def test_export_text(tmp_path):
    model = make_test_model()
    path = export_text(model, tmp_path / "output.txt")
    assert path.exists()
    content = path.read_text()
    assert "OE-001" in content
    assert "PIB Icebreaker" in content
