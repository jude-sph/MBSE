from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from src.models import MBSEModel

# Human-readable display names for known layer keys
_LAYER_DISPLAY_NAMES: dict[str, str] = {
    "operational_analysis": "Operational Analysis",
    "system_analysis": "System Analysis",
    "logical_architecture": "Logical Architecture",
    "physical_architecture": "Physical Architecture",
}


def _layer_display_name(key: str) -> str:
    return _LAYER_DISPLAY_NAMES.get(key, key.replace("_", " ").title())


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------


def export_json(model: MBSEModel, output_path: Path) -> Path:
    """Serialize *model* to a formatted JSON file and return the path."""
    output_path = Path(output_path)
    data = json.loads(model.model_dump_json())
    output_path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    return output_path


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def _write_layer_sheet(wb: openpyxl.Workbook, layer_key: str, layer_data: Any) -> None:
    """Write one worksheet for *layer_data* (a dict of lists of element dicts)."""
    sheet_name = _layer_display_name(layer_key)[:31]  # Excel 31-char limit
    ws = wb.create_sheet(title=sheet_name)

    if not isinstance(layer_data, dict):
        return

    row = 1
    for element_type, elements in layer_data.items():
        if not elements:
            continue

        # Section header
        ws.cell(row=row, column=1, value=element_type.replace("_", " ").title())
        row += 1

        # Collect all keys from all elements of this type
        if isinstance(elements[0], dict):
            headers = list(elements[0].keys())
        else:
            headers = ["value"]

        # Column headers
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1

        # Data rows
        for element in elements:
            if isinstance(element, dict):
                for col, header in enumerate(headers, start=1):
                    value = element.get(header, "")
                    if isinstance(value, list):
                        value = ", ".join(str(v) for v in value)
                    ws.cell(row=row, column=col, value=value)
            else:
                ws.cell(row=row, column=1, value=str(element))
            row += 1

        row += 1  # blank separator between element types


def export_xlsx(model: MBSEModel, output_path: Path) -> Path:
    """Export *model* to an Excel workbook and return the path."""
    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # One sheet per layer
    for layer_key, layer_data in model.layers.items():
        _write_layer_sheet(wb, layer_key, layer_data)

    # Requirements sheet
    ws_req = wb.create_sheet(title="Requirements")
    ws_req.append(["ID", "Text", "Source DIG"])
    for req in model.requirements:
        ws_req.append([req.id, req.text, req.source_dig])

    # Links sheet
    ws_links = wb.create_sheet(title="Links")
    ws_links.append(["Source", "Type", "Target", "Description"])
    for link in model.links:
        ws_links.append([link.source, link.type, link.target, link.description])

    # Instructions sheet
    ws_inst = wb.create_sheet(title="Instructions")
    ws_inst.append(["Step", "Action", "Detail", "Layer"])
    for step in model.instructions.get("steps", []):
        if isinstance(step, dict):
            ws_inst.append([
                step.get("step", ""),
                step.get("action", ""),
                step.get("detail", ""),
                step.get("layer", ""),
            ])

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Plain text
# ---------------------------------------------------------------------------

def export_text(model: MBSEModel, output_path: Path) -> Path:
    """Export *model* as a formatted plain-text hierarchical document."""
    output_path = Path(output_path)
    lines: list[str] = []

    # Header
    tool = model.instructions.get("tool", "Unknown tool")
    timestamp = model.meta.generated_at.strftime("%Y-%m-%d %H:%M:%S UTC") if model.meta.generated_at else datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    lines.append("=" * 72)
    lines.append(f"MBSE Model Export")
    lines.append(f"Mode    : {model.meta.mode}")
    lines.append(f"Tool    : {tool}")
    lines.append(f"Source  : {model.meta.source_file}")
    lines.append(f"Generated: {timestamp}")
    lines.append("=" * 72)
    lines.append("")

    # Layers
    for layer_key, layer_data in model.layers.items():
        display = _layer_display_name(layer_key)
        lines.append(f"## {display}")
        lines.append("-" * 40)

        if isinstance(layer_data, dict):
            for element_type, elements in layer_data.items():
                if not elements:
                    continue
                lines.append(f"  {element_type.replace('_', ' ').title()}:")
                for element in elements:
                    if isinstance(element, dict):
                        elem_id = element.get("id", "")
                        elem_name = element.get("name", "")
                        if elem_id and elem_name:
                            lines.append(f"    [{elem_id}] {elem_name}")
                        elif elem_id:
                            lines.append(f"    [{elem_id}]")
                        elif elem_name:
                            lines.append(f"    {elem_name}")
                        # Additional fields (skip id/name already shown)
                        for k, v in element.items():
                            if k in ("id", "name"):
                                continue
                            if isinstance(v, list):
                                v = ", ".join(str(i) for i in v)
                            lines.append(f"      {k}: {v}")
                    else:
                        lines.append(f"    {element}")
                lines.append("")
        lines.append("")

    # Requirements
    if model.requirements:
        lines.append("## Requirements")
        lines.append("-" * 40)
        for req in model.requirements:
            lines.append(f"  [{req.id}] {req.text}")
            lines.append(f"    Source: {req.source_dig}")
        lines.append("")

    # Links
    if model.links:
        lines.append("## Links")
        lines.append("-" * 40)
        for link in model.links:
            lines.append(f"  [{link.id}] {link.source} --{link.type}--> {link.target}")
            if link.description:
                lines.append(f"    {link.description}")
        lines.append("")

    # Instructions
    steps = model.instructions.get("steps", [])
    if steps:
        lines.append("## Instructions")
        lines.append("-" * 40)
        for step in steps:
            if isinstance(step, dict):
                lines.append(f"  Step {step.get('step', '?')}: {step.get('action', '')}")
                lines.append(f"    {step.get('detail', '')}")
                lines.append(f"    Layer: {step.get('layer', '')}")
        lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")
    return output_path
